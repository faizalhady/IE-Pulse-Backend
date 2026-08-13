"""
api/routers/access.py
─────────────────────
Who can do what, what they lead, and what they get emailed about.

One row per PERSON, not one per grant. A person has:
    level      viewer | admin | super_admin | developer
    apps       'all', or a CSV of module ids
    workcells  zero or many — leading one is a separate fact from level
    notify     opt-ins, keyed ('ole_smh' today)

Deliberately NOT a proxy to AD_GET. That service needs Windows auth and this
backend runs as LocalSystem, so it would authenticate as the machine account
rather than a person. The browser does NTLM for free in the intranet zone, so
the UI searches AD_GET directly and posts whoever it picked.

POSITION AND CUSTOMER ARE THE EXCEPTION — they are resolved HERE, not trusted
from the client. See `_hc_index`.

Endpoints:
  GET    /api/access                users, with workcells + notifications
  GET    /api/access/me/{ntid}      one person's effective access
  GET    /api/access/recipients     who to email for a notification key
  PUT    /api/access/{ntid}         upsert a person (whole record — see below)
  DELETE /api/access/{ntid}         remove them entirely
"""

import logging
import os
from functools import lru_cache
from pathlib import Path
from typing import Literal, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel, Field

from core.auth import require_level, verified_ntid
from core.database import get_conn

log = logging.getLogger(__name__)

router = APIRouter(prefix="/api/access", tags=["Access"])

Level = Literal["viewer", "admin", "super_admin", "developer"]

# Weakest to strongest — used for "is this person at least X?" checks.
_RANK = {"viewer": 0, "admin": 1, "super_admin": 2, "developer": 3}


# ─── Headcount lookup ────────────────────────────────────────────────────────
# position/customer used to come from the client: the UI read them off AD_GET
# when you picked the person and posted them along. That made every browser
# tab a source of truth. A build that omitted the two fields sent no value,
# Pydantic defaulted them to None, and the upsert wrote NULL over whatever was
# there — three people were added that way before anyone noticed, and editing
# an existing person would have wiped theirs too.
#
# HC.xlsx is the file AD_GET itself serves from, and it sits on this box, so
# read it directly. No Windows auth involved — it is a file, not a service.
# Now it does not matter what the client sends, or which bundle version the
# browser is running: the values come from headcount either way.
HC_XLSX = Path(os.getenv("HC_XLSX", r"D:\Application\RetrieveUserInfo\data\HC.xlsx"))

# Spreadsheet column -> what we call it. Header text as AD_GET's HeadcountStore
# reads it; a rename there breaks the lookup loudly (empty index) rather than
# silently writing the wrong person's title.
_HC_COLS = {"NT Account Name": "ntid", "Employee ID": "employee_id",
            "Primary Work Email": "email", "Business Title": "position",
            "Customer": "customer"}


def _hc_mtime() -> int:
    """Cache key — a headcount refresh must invalidate the index."""
    try:
        return HC_XLSX.stat().st_mtime_ns
    except OSError:
        return 0


@lru_cache(maxsize=2)
def _hc_index(_key: int) -> dict[str, tuple[Optional[str], Optional[str]]]:
    """lookup key -> (position, customer). ~12,300 rows, read once per refresh.

    Indexed under NTID, employee id AND email because `user_access.ntid` holds
    both spellings people are known by — numeric ('123755') and account name
    ('LawC2'). All keys are casefolded.
    """
    if not HC_XLSX.exists():
        log.warning("access: headcount not found at %s — position/customer "
                    "will fall back to whatever the client sent", HC_XLSX)
        return {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(HC_XLSX, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
        idx = {_HC_COLS[h]: i for i, h in enumerate(header) if h in _HC_COLS}
        missing = set(_HC_COLS.values()) - set(idx)
        if missing:
            log.error("access: headcount is missing columns %s — not indexing", missing)
            return {}

        out: dict[str, tuple[Optional[str], Optional[str]]] = {}
        for r in rows:
            def cell(name: str) -> Optional[str]:
                v = r[idx[name]]
                return str(v).strip() or None if v is not None else None
            pair = (cell("position"), cell("customer"))
            if pair == (None, None):
                continue
            for key_col in ("ntid", "employee_id", "email"):
                k = cell(key_col)
                if k:
                    out.setdefault(k.casefold(), pair)
        wb.close()
        log.info("access: headcount indexed — %d keys from %s", len(out), HC_XLSX.name)
        return out
    except Exception:                                   # noqa: BLE001
        # Never take the roster down because a spreadsheet moved or is locked.
        log.exception("access: could not read headcount — falling back to client values")
        return {}


def hc_person(key: Optional[str]) -> tuple[Optional[str], Optional[str]]:
    """(position, customer) from headcount, or (None, None) if unknown."""
    if not key:
        return (None, None)
    return _hc_index(_hc_mtime()).get(key.casefold(), (None, None))


def resolve_hc(ntid: Optional[str], email: Optional[str],
               position: Optional[str] = None,
               customer: Optional[str] = None) -> tuple[Optional[str], Optional[str]]:
    """Headcount first, the caller's values only as a fallback.

    Neither field is editable in the UI, so a client value is at best a stale
    copy of this same file. Falling back to it still matters for people
    headcount has never heard of — contractors, or someone added before their
    HC row lands.
    """
    pos, cust = hc_person(ntid)
    if not (pos or cust):                       # NTID unknown — try their email
        pos, cust = hc_person(email)
    return (pos or position, cust or customer)


# One statement, one place. The COALESCEs are the guarantee that a save can
# never destroy a value it did not supply — tests/test_access_hc.py runs THIS
# string, so weakening it here fails the check rather than passing quietly.
UPSERT_SQL = """
    INSERT INTO user_access (ntid, name, email, position, customer, level, apps, added_by)
    VALUES (?,?,?,?,?,?,?,?)
    ON CONFLICT (ntid) DO UPDATE SET
        name     = COALESCE(excluded.name,     user_access.name),
        email    = COALESCE(excluded.email,    user_access.email),
        position = COALESCE(excluded.position, user_access.position),
        customer = COALESCE(excluded.customer, user_access.customer),
        level = excluded.level, apps = excluded.apps,
        updated_at = datetime('now')
"""


class UserIn(BaseModel):
    """A person's WHOLE access record.

    PUT, not PATCH: the edit dialog shows every workcell and app at once, so it
    always knows the complete intended state. Sending the whole thing means
    un-ticking a workcell actually removes it — a partial update would need a
    separate 'remove' call and could silently leave orphans behind.
    """
    name: Optional[str] = Field(None, max_length=200)
    email: Optional[str] = Field(None, max_length=200)
    # From headcount — businessTitle and customer, carried at add time.
    position: Optional[str] = Field(None, max_length=200)
    customer: Optional[str] = Field(None, max_length=200)
    level: Level = "viewer"
    # 'all' rather than enumerating every module, so adding a new app does not
    # silently narrow everyone's existing access.
    apps: list[str] = Field(default_factory=lambda: ["all"])
    workcells: list[str] = Field(default_factory=list)
    primary_workcell: Optional[str] = None
    # key -> enabled, e.g. {"ole_smh": true}
    notifications: dict[str, bool] = Field(default_factory=dict)
    added_by: Optional[str] = Field(None, max_length=40)


def _load_users(conn, ntid: Optional[str] = None) -> list[dict]:
    sql = "SELECT * FROM user_access"
    args: list = []
    if ntid:
        sql += " WHERE lower(ntid) = lower(?)"
        args.append(ntid)
    sql += (" ORDER BY CASE level WHEN 'developer' THEN 0 WHEN 'super_admin' THEN 1 "
            "WHEN 'admin' THEN 2 ELSE 3 END, name COLLATE NOCASE")
    users = [dict(r) for r in conn.execute(sql, args)]
    if not users:
        return []

    ids = [u["ntid"] for u in users]
    marks = ",".join("?" * len(ids))
    wc: dict[str, list] = {}
    for r in conn.execute(
            f"SELECT * FROM user_workcells WHERE ntid IN ({marks}) ORDER BY workcell", ids):
        wc.setdefault(r["ntid"], []).append({"workcell": r["workcell"], "is_primary": r["is_primary"]})
    nt: dict[str, dict] = {}
    for r in conn.execute(
            f"SELECT * FROM user_notifications WHERE ntid IN ({marks})", ids):
        nt.setdefault(r["ntid"], {})[r["key"]] = bool(r["enabled"])

    for u in users:
        u["apps"] = u["apps"].split(",") if u["apps"] else ["all"]
        u["workcells"] = [w["workcell"] for w in wc.get(u["ntid"], [])]
        u["primary_workcell"] = next(
            (w["workcell"] for w in wc.get(u["ntid"], []) if w["is_primary"]), None)
        u["notifications"] = nt.get(u["ntid"], {})
    return users


def _guard_demotion(conn, ntid: str, caller: str, new_level: Optional[str]) -> None:
    """Stop the two ways a developer can lock the roster out of its own admins.

    `new_level=None` means deletion. Demoting yourself is refused outright — the
    endpoint that would undo it needs developer, so the fix is a hand-written
    SQLite UPDATE on the server (which is exactly how 5 Aug went).
    """
    row = conn.execute(
        "SELECT level FROM user_access WHERE lower(ntid) = lower(?)", (ntid,)).fetchone()
    was_dev = bool(row) and row["level"] == "developer"
    if not was_dev or new_level == "developer":
        return
    if ntid.lower() == caller.lower():
        raise HTTPException(403, "You cannot remove your own developer access — ask another developer.")
    n = conn.execute("SELECT count(*) c FROM user_access WHERE level = 'developer'").fetchone()["c"]
    if n <= 1:
        raise HTTPException(403, "This is the last developer — promote someone else first.")


def _heal_missing_hc(conn, users: list[dict]) -> None:
    """Fill any NULL position/customer from headcount, in place and in the DB.

    Self-healing rather than a one-off migration: rows can be NULL because they
    were written by an old client, because HC was unreachable at save time, or
    because the person's HC record only landed later. Opening the page fixes
    all three. Only touches rows that are actually missing something.
    """
    for u in users:
        if u.get("position") and u.get("customer"):
            continue
        pos, cust = resolve_hc(u["ntid"], u.get("email"),
                               u.get("position"), u.get("customer"))
        if (pos, cust) == (u.get("position"), u.get("customer")):
            continue                                    # headcount knows no more
        conn.execute("UPDATE user_access SET position = ?, customer = ? WHERE ntid = ?",
                     (pos, cust, u["ntid"]))
        u["position"], u["customer"] = pos, cust
        log.info("access: filled %s from headcount — %s / %s", u["ntid"], pos, cust)


@router.get("", dependencies=[Depends(verified_ntid)])
def list_users():
    with get_conn() as conn:
        users = _load_users(conn)
        _heal_missing_hc(conn, users)
    return {"count": len(users), "users": users}


@router.get("/me/{ntid}")
def effective_access(ntid: str, caller: str = Depends(verified_ntid)):
    """What this person may do — the shape a permission check needs.

    `all_workcells` is true for super_admin and developer: enumerating every
    workcell for them would go stale the moment one is added, and "can edit
    anything" is the actual intent.
    """
    if ntid.lower() != caller.lower():
        with get_conn() as conn:
            row = conn.execute(
                "SELECT level FROM user_access WHERE lower(ntid) = lower(?)", (caller,)
            ).fetchone()
        if not row or row["level"] not in ("admin", "super_admin", "developer"):
            raise HTTPException(403, "You can only look up your own access.")

    with get_conn() as conn:
        users = _load_users(conn, ntid)
    if not users:
        # Absence is not an error — everyone who was never granted anything is
        # a viewer. Returning 404 would make every caller special-case it.
        return {"ntid": ntid, "known": False, "level": "viewer", "apps": ["all"],
                "workcells": [], "all_workcells": False, "is_admin": False,
                "notifications": {}}
    u = users[0]
    return {
        "ntid": u["ntid"], "known": True, "level": u["level"], "apps": u["apps"],
        "workcells": u["workcells"],
        "all_workcells": _RANK[u["level"]] >= _RANK["super_admin"],
        "is_admin": _RANK[u["level"]] >= _RANK["admin"],
        "notifications": u["notifications"],
        "name": u["name"], "email": u["email"],
        "position": u["position"], "customer": u["customer"],
    }


@router.get("/recipients", dependencies=[Depends(verified_ntid)])
def recipients(key: str = Query("ole_smh", description="Notification key, e.g. 'ole_smh'.")):
    """Who to email for a notification, grouped by the workcell it concerns.

    People opted in but with no email in headcount come back under
    `unreachable` rather than being dropped — a missing recipient otherwise
    looks exactly like "nothing to report".
    """
    with get_conn() as conn:
        rows = [dict(r) for r in conn.execute("""
            SELECT ua.ntid, ua.name, ua.email, uw.workcell, uw.is_primary
            FROM user_notifications un
            JOIN user_access    ua ON ua.ntid = un.ntid
            JOIN user_workcells uw ON uw.ntid = un.ntid
            WHERE un.key = ? AND un.enabled = 1
            ORDER BY uw.workcell COLLATE NOCASE, uw.is_primary DESC
        """, (key,))]

    by_wc: dict[str, dict] = {}
    unreachable: list[dict] = []
    for r in rows:
        if not r["email"]:
            unreachable.append({"workcell": r["workcell"], "ntid": r["ntid"], "name": r["name"]})
            continue
        w = by_wc.setdefault(r["workcell"], {"workcell": r["workcell"], "to": [], "cc": []})
        # Primary in To, the rest Cc — one owner, everyone else informed.
        (w["to"] if r["is_primary"] else w["cc"]).append(
            {"ntid": r["ntid"], "name": r["name"], "email": r["email"]})

    for w in by_wc.values():
        if not w["to"] and w["cc"]:
            w["to"].append(w["cc"].pop(0))     # never leave To empty

    return {
        "key": key,
        "count": len(by_wc),
        "workcells": sorted(by_wc.values(), key=lambda x: x["workcell"].lower()),
        "unreachable": unreachable,
    }


@router.put("/{ntid}")
def upsert_user(ntid: str, body: UserIn, caller: str = Depends(require_level("developer"))):
    apps = ",".join(body.apps) if body.apps else "all"

    position, customer = resolve_hc(ntid, body.email, body.position, body.customer)

    with get_conn() as conn:
        _guard_demotion(conn, ntid, caller, body.level)
        conn.execute(UPSERT_SQL, (ntid, body.name, body.email, position, customer,
                                  body.level, apps, body.added_by))

        # Replace rather than merge: the dialog always sends the full intended
        # set, so anything absent was deliberately un-ticked.
        conn.execute("DELETE FROM user_workcells WHERE ntid = ?", (ntid,))
        for w in body.workcells:
            conn.execute(
                "INSERT INTO user_workcells (ntid, workcell, is_primary) VALUES (?,?,?)",
                (ntid, w, int(w == body.primary_workcell)))

        conn.execute("DELETE FROM user_notifications WHERE ntid = ?", (ntid,))
        for k, on in body.notifications.items():
            if on:      # only store opt-INS; absence means off
                conn.execute(
                    "INSERT INTO user_notifications (ntid, key, enabled) VALUES (?,?,1)", (ntid, k))

        user = _load_users(conn, ntid)[0]

    log.info("access: %s = %s, apps=%s, workcells=%s", body.name or ntid,
             body.level, apps, body.workcells or "none")
    return user


@router.delete("/{ntid}")
def delete_user(ntid: str, caller: str = Depends(require_level("developer"))):
    with get_conn() as conn:
        _guard_demotion(conn, ntid, caller, None)
        cur = conn.execute("DELETE FROM user_access WHERE lower(ntid) = lower(?)", (ntid,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail=f"No access record for {ntid}.")
        conn.execute("DELETE FROM user_workcells WHERE ntid = ?", (ntid,))
        conn.execute("DELETE FROM user_notifications WHERE ntid = ?", (ntid,))
    return {"deleted": ntid}
